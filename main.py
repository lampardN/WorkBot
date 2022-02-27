import telebot
import os
import openpyxl
import sqlite3
import time
from multiprocessing.context import Process
import schedule
import calendar
import datetime as dt


token = "5115667742:AAGZ8QiWAuRDMXLI2lwvnsTgXAruGYo-Wqg"

bot = telebot.TeleBot(token)


@bot.message_handler(commands=['start'])
def start(message):
    chat_id = message.chat.id
    bot.send_message(chat_id, """Для получения справки введите /help""")
    db = sqlite3.connect("users.db")
    cur = db.cursor()
    cur.execute("select * from users where chat_id=:chat_id", {"chat_id": chat_id})
    if not cur.fetchall():
        cur.execute("insert into users values (?)", (chat_id,))
        db.commit()
    db.close()


@bot.message_handler(commands=['help'])
def Help(message):

    bot.send_message(message.chat.id, """На данном этапе бот поддерживает выгрузку файлов с графиком практик, где все практики пронумерованы и после номера обязательно стоит точка
Так же каждый день в 9 утра будут высылаться сообщения-напоминания""")


@bot.message_handler(commands=['remind'])
def remind(message):
    messageToUser(message.chat.id)


@bot.message_handler(commands=['test'])
def test(message):
    messageToUser()


@bot.message_handler(content_types=['document'])
def test(message):
    try:
        """Saving file"""
        save_dir = "."
        file_name = message.document.file_name
        file_id_info = bot.get_file(message.document.file_id)
        downloaded_file = bot.download_file(file_id_info.file_path)
        src = file_name
        with open(save_dir + "/" + src, 'wb') as new_file:
            new_file.write(downloaded_file)

        """Reading file"""
        wb = openpyxl.load_workbook(filename=file_name)
        sheet = wb[wb.get_sheet_names()[0]]
        data = {}
        increment = 0
        try:
            db = sqlite3.connect("data.db")
        except:
            bot.send_message(message.chat.id, "Не могу выполнить ваш запрос, база открыта другим пользователем")
            return
        cur = db.cursor()
        for row in range(1, sheet.max_row + 1):
            tmp = sheet.cell(row=row, column=1).value
            if tmp is None or not ("1" <= str(tmp)[0] <= "9"):
                continue
            increment += 1
            data[increment] = {}
            data[increment]["group"] = str(sheet.cell(row=row, column=6).value)
            data[increment]["practice_name"] = str(sheet.cell(row=row, column=7).value)
            data[increment]["begin"] = str(sheet.cell(row=row, column=8).value.date().strftime("%d.%m.%Y"))
            data[increment]["end"] = str(sheet.cell(row=row, column=9).value.strftime("%d.%m.%Y"))
            cur.execute(
                """select * from orders where \"group\"=:group and practice_name=:practice_name and begin=:begin and end=:end""",
                data[increment])
            if not cur.fetchall():
                cur.execute("""insert into orders(\"group\", practice_name, begin, end, status) values (?, ?, ?, ?)""",
                            (data[increment]["group"],
                             data[increment]["practice_name"],
                             data[increment]["begin"],
                             data[increment]["end"]))
                db.commit()

        db.close()
        bot.send_message(message.chat.id, "Информация успешно добавлена в базу")
        path = os.path.join(os.path.abspath(os.path.dirname(__file__)), file_name)
        os.remove(path)
    except:
        bot.send_message(message.chat.id, "Не могу выполнить ваш запрос")


class ScheduleMessage():
    def try_send_schedule():
        while True:
            schedule.run_pending()
            time.sleep(1)

    def start_process():
        p1 = Process(target=ScheduleMessage.try_send_schedule, args=())
        p1.start()


@bot.callback_query_handler(func=lambda call: True)
def button_query(call):
    identifier = int(call.data)
    try:
        db = sqlite3.connect("data.db")
        db.cursor().execute("delete from orders where id=:id", {"id": identifier})
        db.commit()
        bot.answer_callback_query(callback_query_id=call.id, show_alert=True, text="Запись успешно удалена")
        db.close()
    except:
        bot.answer_callback_query(callback_query_id=call.id, show_alert=True, text="Запись удалена из ваших заметок")
    bot.delete_message(call.message.chat.id, call.message.message_id)


def messageToUser(users=None):
    if users is None:
        db = sqlite3.connect("users.db")
        users = db.cursor().execute("select * from users").fetchall()
        db.close()
    else:
        users = [(users,)]
    today = dt.date.today()
    days = calendar.monthrange(today.year, today.month)[1]
    next_month_date = today + dt.timedelta(days=days)
    next_month_date = dt.datetime.strptime(next_month_date.strftime("%d.%m.%Y"), "%d.%m.%Y")

    db = sqlite3.connect("data.db")
    data = db.cursor().execute("select * from orders").fetchall()
    db.close()
    data = [x for x in data if dt.datetime.strptime(x[3], "%d.%m.%Y") <= next_month_date]
    for user in users:
        for entry in data:
            try:
                markup = telebot.types.InlineKeyboardMarkup()
                btn = telebot.types.InlineKeyboardButton(text="Сделано", callback_data=f"{entry[0]}")
                markup.add(btn)
                bot.send_message(user[0], f"""Группа: {entry[1]}\nНазвание: {entry[2]}\nНачало: {entry[3]}\nКонец: {entry[4]}""", reply_markup=markup)
            except:
                pass


def send_message():
    messageToUser()


schedule.every().day.at("01:00").do(send_message)


if __name__ == "__main__":
    ScheduleMessage.start_process()
    try:
        bot.infinity_polling()
    except:
        pass
